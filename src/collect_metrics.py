"""
Скрипт для сбора метрик из metrics_summary.txt файлов
и создания сводной таблицы в Excel и CSV форматах.
"""

import os
import re
import argparse
import csv
from pathlib import Path

# Попробуем импортировать openpyxl для Excel (опционально)
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def parse_metrics_file(filepath: Path) -> dict:
    """Парсит txt файл с метриками вида 'Key: value ± std'"""
    metrics = {}
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                # Формат: "Duration Accuracy:  1.000 ± 0.000"
                match = re.match(r"^(.+?):\s+([\d.]+)\s*±\s*([\d.]+)", line)
                if match:
                    key = match.group(1).strip()
                    value = float(match.group(2))
                    std = float(match.group(3))
                    metrics[key] = {"value": value, "std": std}
                else:
                    # Формат без std: "Key: value"
                    match2 = re.match(r"^(.+?):\s+([\d.]+)", line)
                    if match2:
                        key = match2.group(1).strip()
                        metrics[key] = {"value": float(match2.group(2)), "std": None}
    except Exception as e:
        print(f"  [!] Ошибка при чтении {filepath}: {e}")
    return metrics


def collect_all_metrics(root_dir: Path, txt_filename: str = "metrics_summary.txt") -> list[dict]:
    """Рекурсивно обходит папки и собирает метрики."""
    rows = []
    found = sorted(root_dir.rglob(txt_filename))

    if not found:
        print(f"[!] Файлы '{txt_filename}' не найдены в {root_dir}")
        return rows

    print(f"Найдено файлов: {len(found)}\n")

    for filepath in found:
        # Имя задачи = имя папки, содержащей txt файл
        task_name = filepath.parent.name
        rel_path = filepath.relative_to(root_dir)
        print(f"  Обрабатываю: {rel_path}")

        metrics = parse_metrics_file(filepath)
        if metrics:
            row = {"task": task_name, "path": str(rel_path.parent)}
            for metric_name, vals in metrics.items():
                row[metric_name] = vals["value"]
                if vals["std"] is not None:
                    row[f"{metric_name} ±"] = vals["std"]
            rows.append(row)

    return rows


def save_csv(rows: list[dict], out_path: Path):
    """Сохраняет результаты в CSV."""
    if not rows:
        return

    # Собираем все уникальные колонки, сохраняя порядок
    all_keys = ["task", "path"]
    for row in rows:
        for k in row:
            if k not in all_keys:
                all_keys.append(k)

    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=all_keys, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in all_keys})

    print(f"\n✓ CSV сохранён: {out_path}")


def save_excel(rows: list[dict], out_path: Path):
    """Сохраняет красивую Excel таблицу."""
    if not rows:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Metrics Summary"

    # Собираем все колонки
    all_keys = ["task", "path"]
    for row in rows:
        for k in row:
            if k not in all_keys:
                all_keys.append(k)

    # Стили
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill("solid", fgColor="D6E4F0")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Заголовки
    for col_idx, key in enumerate(all_keys, 1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Данные
    for row_idx, row in enumerate(rows, 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, key in enumerate(all_keys, 1):
            val = row.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = center if col_idx > 2 else left
            if fill:
                cell.fill = fill
            # Числа: 3 знака после запятой
            if isinstance(val, float):
                cell.number_format = "0.000"

    # Автоширина колонок
    for col_idx, key in enumerate(all_keys, 1):
        max_len = max(len(str(key)), *(len(str(row.get(key, ""))) for row in rows))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # Заморозить первую строку
    ws.freeze_panes = "A2"

    # Автофильтр
    ws.auto_filter.ref = ws.dimensions

    wb.save(out_path)
    print(f"✓ Excel сохранён: {out_path}")


def print_table(rows: list[dict]):
    """Выводит простую таблицу в терминал."""
    if not rows:
        print("Нет данных.")
        return

    all_keys = ["task"]
    for row in rows:
        for k in row:
            if k not in all_keys and k != "path":
                all_keys.append(k)

    # Ширины колонок
    widths = {k: max(len(k), *(len(f"{row.get(k, ''):.3f}" if isinstance(row.get(k), float) else str(row.get(k, ""))) for row in rows)) for k in all_keys}

    sep = "+" + "+".join("-" * (widths[k] + 2) for k in all_keys) + "+"
    header = "|" + "|".join(f" {k.center(widths[k])} " for k in all_keys) + "|"

    print("\n" + sep)
    print(header)
    print(sep)
    for row in rows:
        line = "|"
        for k in all_keys:
            val = row.get(k, "")
            if isinstance(val, float):
                val = f"{val:.3f}"
            else:
                val = str(val)
            line += f" {val.ljust(widths[k])} |"
        print(line)
    print(sep)


def main():
    parser = argparse.ArgumentParser(description="Сбор метрик из txt файлов в сводную таблицу")
    parser.add_argument("--root", default=".", help="Корневая папка для поиска (по умолчанию: текущая)")
    parser.add_argument("--filename", default="metrics_summary.txt", help="Имя файла с метриками")
    parser.add_argument("--out", default="metrics_summary", help="Имя выходного файла (без расширения)")
    args = parser.parse_args()

    root = Path(args.root)
    if not root.exists():
        print(f"[!] Папка не найдена: {root}")
        return

    print(f"Сканирую: {root.resolve()}\n")
    rows = collect_all_metrics(root, args.filename)

    if not rows:
        print("Метрики не найдены.")
        return

    print_table(rows)

    out_dir = root  # Сохраняем рядом с данными
    save_csv(rows, out_dir / f"{args.out}.csv")

    if EXCEL_AVAILABLE:
        save_excel(rows, out_dir / f"{args.out}.xlsx")
    else:
        print("\n[i] openpyxl не установлен — Excel не создаётся.")
        print("    Установить: pip install openpyxl")


if __name__ == "__main__":
    main()